from csv import reader
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import gettempdir
from time import time
from uuid import uuid4

from openpyxl import load_workbook


MAX_FILE_SIZE = 25 * 1024 * 1024
MAX_ROWS = 50_000
MAX_COLUMNS = 100
SUPPORTED_EXTENSIONS = {'.csv', '.xlsx'}

IMPORT_CACHE_DIR = Path(gettempdir()) / 'dreitrack_imports'
IMPORT_CACHE_MAX_AGE = 60 * 60

IMPORT_FIELDS = {
    'sku': 'SKU',
    'name': 'Item Name',
    'quantity': 'Initial Quantity',
    'category': 'Category',
    'storage_location': 'Storage Location',
    'supplier': 'Supplier',
    'unit_cost': 'Unit Cost',
    'minimum_stock': 'Minimum Stock',
    'lead_time_days': 'Lead Time Days',
}

FIELD_ALIASES = {
    'sku': {
        'sku',
        'part no',
        'part number',
        'item code',
        'stock code',
        'material code',
        'product code',
    },
    'name': {
        'name',
        'item name',
        'description',
        'item description',
        'product name',
    },
    'quantity': {
        'quantity',
        'qty',
        'stock',
        'stock quantity',
        'on hand',
        'on hand quantity',
    },
    'category': {
        'category',
        'item category',
        'product category',
    },
    'storage_location': {
        'storage location',
        'location',
        'shelf',
        'bin',
        'rack',
    },
    'supplier': {
        'supplier',
        'vendor',
        'supplier name',
        'vendor name',
    },
    'unit_cost': {
        'unit cost',
        'cost',
        'unit price',
        'price',
    },
    'minimum_stock': {
        'minimum stock',
        'min stock',
        'minimum quantity',
        'reorder level',
    },
    'lead_time_days': {
        'lead time',
        'lead time days',
        'lead days',
    },
}


def parse_inventory_file(
    filename: str,
    content: bytes,
) -> tuple[list[str], list[list[str]]]:
    if len(content) > MAX_FILE_SIZE:
        raise ValueError('Import file is larger than 25 MB.')

    extension = Path(filename).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError('Only .csv and .xlsx inventory files are supported.')

    rows = _read_csv(content) if extension == '.csv' else _read_xlsx(content)
    rows = _remove_empty_rows(rows)

    if not rows:
        raise ValueError('The import file does not contain any inventory data.')

    headers = rows[0]
    data = rows[1:]

    if not any(headers):
        raise ValueError('The first row must contain column names.')

    if not data:
        raise ValueError('The import file contains headers but no inventory rows.')

    return headers, data


def _read_csv(content: bytes) -> list[list[str]]:
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ValueError('CSV files must use UTF-8 encoding.') from exc

    return _read_rows(reader(StringIO(text)))


def _read_xlsx(content: bytes) -> list[list[str]]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError('The Excel file could not be read.') from exc

    try:
        return _read_rows(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def _read_rows(rows) -> list[list[str]]:
    result = []

    for row_number, row in enumerate(rows, start=1):
        if row_number > MAX_ROWS + 1:
            raise ValueError(
                f'Import files cannot contain more than {MAX_ROWS:,} data rows.'
            )

        if len(row) > MAX_COLUMNS:
            raise ValueError(
                f'Import files cannot contain more than {MAX_COLUMNS} columns.'
            )

        result.append([_clean_value(value) for value in row])

    return result


def _clean_value(value) -> str:
    return '' if value is None else str(value).strip()


def _remove_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    return [row for row in rows if any(row)]


def suggest_column_mapping(headers: list[str]) -> dict[str, str]:
    mapping = {}

    for header in headers:
        normalized = _normalize_header(header)

        for field, aliases in FIELD_ALIASES.items():
            if normalized in aliases:
                mapping[header] = field
                break

    return mapping


def _normalize_header(value: str) -> str:
    return ' '.join(
        value.lower()
        .replace('_', ' ')
        .replace('-', ' ')
        .replace('.', ' ')
        .split()
    )


def store_import_file(user_id: int, filename: str, content: bytes) -> str:
    _cleanup_import_cache()

    extension = Path(filename).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError('Only .csv and .xlsx inventory files are supported.')

    token = uuid4().hex
    IMPORT_CACHE_DIR.mkdir(parents=True, exist_ok=True)

    path = IMPORT_CACHE_DIR / f'{user_id}_{token}{extension}'
    path.write_bytes(content)

    return token


def load_import_file(user_id: int, token: str) -> tuple[str, bytes]:
    _cleanup_import_cache()

    if len(token) != 32 or any(
        character not in '0123456789abcdef'
        for character in token
    ):
        raise ValueError('Invalid inventory import token.')

    for extension in SUPPORTED_EXTENSIONS:
        path = IMPORT_CACHE_DIR / f'{user_id}_{token}{extension}'

        if path.exists():
            return f'import{extension}', path.read_bytes()

    raise ValueError(
        'This inventory import preview has expired. Please upload the file again.'
    )


def delete_import_file(user_id: int, token: str) -> None:
    for extension in SUPPORTED_EXTENSIONS:
        path = IMPORT_CACHE_DIR / f'{user_id}_{token}{extension}'

        if path.exists():
            path.unlink()


def _cleanup_import_cache() -> None:
    if not IMPORT_CACHE_DIR.exists():
        return

    cutoff = time() - IMPORT_CACHE_MAX_AGE

    for path in IMPORT_CACHE_DIR.iterdir():
        if path.is_file() and path.stat().st_mtime < cutoff:
            path.unlink()


def validate_inventory_rows(
    rows: list[list[str]],
    mapping: dict[int, str],
    existing_skus: set[str],
    default_category: str,
    default_location_id: int,
    location_lookup: dict[str, int],
) -> tuple[list[dict], list[str]]:
    validated = []
    errors = []
    file_skus = set()

    for row_number, row in enumerate(rows, start=2):
        values = {
            field: row[index].strip() if index < len(row) else ''
            for index, field in mapping.items()
        }

        sku = values.get('sku', '')
        name = values.get('name', '')
        category = values.get('category', '') or default_category

        if not sku:
            errors.append(f'Row {row_number}: SKU is required.')

        if not name:
            errors.append(f'Row {row_number}: Item Name is required.')

        if not category:
            errors.append(f'Row {row_number}: Category is required.')

        if sku in existing_skus:
            errors.append(f'Row {row_number}: SKU {sku} already exists.')

        if sku in file_skus:
            errors.append(
                f'Row {row_number}: SKU {sku} appears more than once in the file.'
            )

        if sku:
            file_skus.add(sku)

        quantity = _parse_non_negative_whole_number(
            values.get('quantity', ''),
            'Initial Quantity',
            row_number,
            0,
            errors,
        )

        minimum_stock = _parse_non_negative_whole_number(
            values.get('minimum_stock', ''),
            'Minimum Stock',
            row_number,
            0,
            errors,
        )

        lead_time_days = _parse_non_negative_whole_number(
            values.get('lead_time_days', ''),
            'Lead Time Days',
            row_number,
            14,
            errors,
        )

        unit_cost = _parse_non_negative_number(
            values.get('unit_cost', ''),
            'Unit Cost',
            row_number,
            errors,
        )

        location_id = default_location_id
        location_value = values.get('storage_location', '')

        if location_value:
            location_id = location_lookup.get(location_value.casefold())

            if location_id is None:
                errors.append(
                    f'Row {row_number}: Storage Location '
                    f'"{location_value}" does not exist.'
                )

        validated.append({
            'sku': sku,
            'name': name,
            'category': category,
            'quantity': quantity,
            'storage_location_id': location_id,
            'supplier': values.get('supplier', '') or None,
            'unit_cost': unit_cost,
            'minimum_stock': minimum_stock,
            'lead_time_days': lead_time_days,
        })

    return validated, errors


def _parse_non_negative_whole_number(
    value: str,
    label: str,
    row_number: int,
    default: int,
    errors: list[str],
) -> int | None:
    if not value:
        return default

    try:
        number = Decimal(value)
    except InvalidOperation:
        errors.append(f'Row {row_number}: {label} must be a whole number.')
        return None

    if number < 0 or number != number.to_integral_value():
        errors.append(
            f'Row {row_number}: {label} must be a non-negative whole number.'
        )
        return None

    return int(number)


def _parse_non_negative_number(
    value: str,
    label: str,
    row_number: int,
    errors: list[str],
) -> float | None:
    if not value:
        return None

    try:
        number = Decimal(value)
    except InvalidOperation:
        errors.append(f'Row {row_number}: {label} must be a valid number.')
        return None

    if number < 0:
        errors.append(f'Row {row_number}: {label} cannot be negative.')
        return None

    return float(number)
